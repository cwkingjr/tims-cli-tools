import pandas as pd
from . import file_utils

from openpyxl.styles import Alignment, PatternFill
from openpyxl import load_workbook


def write_consolidated_spreadsheet(
    *,
    original_df: pd.DataFrame,
    all_payments_df: pd.DataFrame,
    all_payments_totals_df: pd.DataFrame,
    payee_tuples,
) -> None:
    """Creates a consolidated spreadsheet.

    Spreadsheet will be dumped to documents folder and contain the original input data on one worksheet,
    all the payments table data on one worksheet, and an individual worksheet for each crew
    member with payments in the pay_to column of the database.
    """
    file_path = file_utils.create_new_spreadsheet_central_tz_filepath(
        filename_prefix="payroll_consolidated"
    )

    underlined_payees = []

    with pd.ExcelWriter(file_path, engine="xlsxwriter") as writer:
        original_df.to_excel(writer, sheet_name="Original_Data", index=False)
        all_payments_totals_df.to_excel(
            writer, sheet_name="All_Payments_Totals", index=False
        )
        all_payments_df.to_excel(writer, sheet_name="All_Payments", index=False)
        for one_payee_tuple in payee_tuples:
            payee, df = one_payee_tuple
            underlined_payee = payee.replace(" ", "_")
            underlined_payees.append(underlined_payee)
            df.to_excel(writer, sheet_name=underlined_payee, index=False)

    format_existing_consolidated_spreadsheet(
        file_path=file_path, payees=underlined_payees
    )


def write_individual_spreadsheet(
    *,
    pay_to_name: str,
    individual_payments_totals_df: pd.DataFrame,
    individual_payments_df: pd.DataFrame,
) -> None:
    """Creates a single individual's spreadsheet for their own record."""
    clean_payto = pay_to_name.lower().replace(" ", "_")
    file_path = file_utils.create_new_spreadsheet_central_tz_filepath(
        filename_prefix=f"payroll_{clean_payto}"
    )

    with pd.ExcelWriter(file_path, engine="xlsxwriter") as writer:
        individual_payments_totals_df.to_excel(
            writer, sheet_name="Payment_Totals", index=False
        )
        individual_payments_df.to_excel(writer, sheet_name="Payments", index=False)

    format_existing_individual_spreadsheet(file_path=file_path)


def format_existing_individual_spreadsheet(*, file_path):
    """Workaround for lack of format access when using df.to_excel."""
    # currency_format = numbers.FORMAT_CURRENCY_USD_SIMPLE  # Example: $#,##0.00
    currency_format = "#,##0.00"
    header_fill = PatternFill(start_color="FF03BDFC", fill_type="solid")

    # Load the existing workbook
    workbook = load_workbook(file_path)

    # Format the Payments worksheet

    payments = workbook["Payments"]
    for cell in payments[1]:  # Iterate through cells in the first row
        cell.fill = header_fill
    for column_letter in "GHIJKLMNO":
        for cell in payments[column_letter]:
            cell.number_format = currency_format
    for column_letter in "ABCDEFGHIJKLMNOPQ":
        payments.column_dimensions[column_letter].width = 15
        for cell in payments[column_letter]:
            cell.alignment = Alignment(horizontal="center", vertical="center")
    # format inspection_date
    for cell in payments["P"]:
        cell.number_format = "hh:mm"  # Example: 14:45
    for cell in payments["C"]:
        cell.number_format = "yyyy-mm-dd"

    # Format the Payment_Totals worksheet
    payment_totals = workbook["Payment_Totals"]
    for cell in payment_totals[1]:  # Iterate through cells in the first row
        cell.fill = header_fill
    for column_letter in "ABCDEFGHIJK":
        payment_totals.column_dimensions[column_letter].width = 15
        for cell in payment_totals[column_letter]:
            cell.alignment = Alignment(horizontal="center", vertical="center")
    for column_letter in "BCDEFGHIJ":
        for cell in payment_totals[column_letter]:
            cell.number_format = currency_format

    workbook.save(file_path)


def format_existing_consolidated_spreadsheet(*, file_path, payees: list[str]):  # noqa: PLR0912
    """Workaround for lack of format access when using df.to_excel."""
    # currency_format = numbers.FORMAT_CURRENCY_USD_SIMPLE  # Example: $#,##0.00
    currency_format = "#,##0.00"
    header_fill = PatternFill(start_color="FF03BDFC", fill_type="solid")

    # Load the existing workbook
    workbook = load_workbook(file_path)

    # Format Original_Data worksheet

    original = workbook["Original_Data"]
    for cell in original[1]:  # Iterate through cells in the first row
        cell.fill = header_fill
    for column_letter in "BCDEFGHIJKLMNOPQ":
        original.column_dimensions[column_letter].width = 15
        for cell in original[column_letter]:
            cell.alignment = Alignment(horizontal="center", vertical="center")
    for cell in original["B"]:
        cell.number_format = "yyyy-mm-dd"

    # Format All_Payments_Totals worksheet

    apt = workbook["All_Payments_Totals"]
    for cell in apt[1]:  # Iterate through cells in the first row
        cell.fill = header_fill
    for column_letter in "ABCDEFGHIJK":
        apt.column_dimensions[column_letter].width = 15
        for cell in apt[column_letter]:
            cell.alignment = Alignment(horizontal="center", vertical="center")
    for column_letter in "BCDEFGHIJ":
        for cell in apt[column_letter]:
            cell.number_format = currency_format

    # Format the All_Payments worksheet
    payments = workbook["All_Payments"]
    for cell in payments[1]:  # Iterate through cells in the first row
        cell.fill = header_fill
    for column_letter in "GHIJKLMNO":
        for cell in payments[column_letter]:
            cell.number_format = currency_format
    for column_letter in "ABCDEFGHIJKLMNOPQ":
        payments.column_dimensions[column_letter].width = 15
        for cell in payments[column_letter]:
            cell.alignment = Alignment(horizontal="center", vertical="center")
    for cell in payments["P"]:
        cell.number_format = "hh:mm"  # Example: 14:45
    for cell in payments["C"]:
        cell.number_format = "yyyy-mm-dd"

    # Format the individual payee pages

    for payee in payees:
        ind = workbook[payee]
        for cell in ind[1]:  # Iterate through cells in the first row
            cell.fill = header_fill
        for column_letter in "GHIJKLMNO":
            for cell in ind[column_letter]:
                cell.number_format = currency_format
        for column_letter in "ABCDEFGHIJKLMNOPQ":
            ind.column_dimensions[column_letter].width = 15
            for cell in ind[column_letter]:
                cell.alignment = Alignment(horizontal="center", vertical="center")
        for cell in ind["P"]:
            cell.number_format = "hh:mm"  # Example: 14:45
        for cell in ind["C"]:
            cell.number_format = "yyyy-mm-dd"

    workbook.save(file_path)
